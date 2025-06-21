import os
import win32com.client as win32
import time
import logging
import shutil
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from utility.send_outlook_email import send_outlook_email
import yaml

with open('updating_config.yaml', 'r', encoding='utf-8') as f:
    USER_CONFIG = yaml.safe_load(f)

if os.environ.get("BASE_DIRECTORY_FROM_MONITOR"):
    USER_CONFIG["base_directory"] = os.environ["BASE_DIRECTORY_FROM_MONITOR"]

TO_RECIPIENTS = USER_CONFIG["email_recipients"]["to"]
CC_RECIPIENTS = USER_CONFIG["email_recipients"]["cc"]
BCC_RECIPIENTS = USER_CONFIG["email_recipients"]["bcc"]
EMAIL_SUBJECT = USER_CONFIG["email_subject_prefix"]

# Global variable: logger instance
logger = None

class ExcelAutomationError(Exception):
    """Custom exception class for Excel automation errors"""
    pass

def setup_logging():
    """Setup logging system with both file and console output"""
    global logger
    
    # Ensure log directory exists
    log_dir = USER_CONFIG["log_directory"]
    Path(log_dir).mkdir(parents=True, exist_ok=True)  # Create log directory if not exists
    
    # Generate meaningful log filename
    current_time = datetime.now()
    log_filename = f"ExcelAutoRefresh_{current_time.strftime('%Y%m%d_%H%M%S')}.log"
    log_filepath = os.path.join(log_dir, log_filename)
    os.environ["log_filepath"] = log_filepath 
    print(log_filepath)
 
    
    # Clear existing handlers to avoid duplication
    logger = logging.getLogger('ExcelAutomation')
    logger.handlers.clear()
    
    # Setup file log handler with UTF-8 encoding
    file_handler = logging.FileHandler(log_filepath, encoding='utf-8')
    file_formatter = logging.Formatter('%(asctime)s | %(levelname)-8s | %(message)s')
    file_handler.setFormatter(file_formatter)
    
    # Setup console log handler
    console_handler = logging.StreamHandler()
    console_formatter = logging.Formatter('[%(asctime)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    console_handler.setFormatter(console_formatter)
    
    # Configure logger
    logger.setLevel(logging.INFO)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    logger.propagate = False  # Prevent propagation
    
    # Log startup information - moved after logger initialization
    logger.info("="*80)
    logger.info("🚀 Excel automation program started")
    logger.info(f"📁 Log directory created/confirmed: {log_dir}")  # This will now have proper formatting
    logger.info(f"📄 Log file: {log_filename}")
    logger.info(f"📁 Configured base directory: {USER_CONFIG['base_directory']}")
    logger.info(f"🏷️ Configured file prefixes: {list(USER_CONFIG['file_configs'].keys())}")
    logger.info("="*80)
    
    return logger

def console_print(message, level='info'):
    """Unified message output function for both console and log file"""
    if logger is None:
        # If logger not initialized yet, output to console with timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] {message}")
        return
    
    # Handle empty string case to ensure timestamp inclusion
    if message == "":
        message = " "  # Replace with space to ensure logger processing
    
    # Use logger for unified processing
    if level.lower() == 'info':
        logger.info(message)
    elif level.lower() == 'warning':
        logger.warning(message)
    elif level.lower() == 'error':
        logger.error(message)
    else:
        logger.info(message)

def safe_execute(func, *args, **kwargs):
    """Safely execute function with retry mechanism"""
    max_retries = USER_CONFIG["advanced_settings"]["max_retries"]  # Get maximum retry count
    retry_delay_base = USER_CONFIG["advanced_settings"]["retry_delay_base"]  # Get retry delay base
    
    # Perform multiple retry attempts
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)  # Execute the passed function
        except Exception as e:
            if attempt < max_retries - 1:  # If retry attempts remaining
                console_print(f"Operation failed, retrying ({attempt + 1}/{max_retries}): {str(e)}", level='warning')
                time.sleep(retry_delay_base ** attempt)  # Exponential backoff wait
            else:
                raise e  # Raise exception on final attempt failure
"""
def get_last_save_author_improved(file_path, has_password=False, open_password=None, workbook_obj=None):
    """
    Improved version: Get Excel file's last author information using multiple methods
    Priority: 1. From opened workbook object 2. From openpyxl 3. Return default message
    """
    # Method 1: If workbook object is provided, get author from opened workbook
    if workbook_obj is not None:
        try:
            # Get author information from opened Excel workbook object
            builtin_props = workbook_obj.BuiltinDocumentProperties  # Get built-in document properties
            last_author = builtin_props("Last Author").Value  # Get last author property
            return last_author if last_author else "Last author info not found from opened workbook"
        except Exception as e:
            console_print(f"Cannot get author info from opened workbook: {str(e)}", level='warning')
            # Continue to next method if failed
    
    # Method 2: Use openpyxl for files without password
    if not has_password:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)  # Open workbook in read-only mode
            last_author = wb.properties.lastModifiedBy  # Get last modified by
            wb.close()  # Close workbook
            return last_author if last_author else "Last author info not found from openpyxl"
        except Exception as e:
            console_print(f"Cannot get author info via openpyxl: {str(e)}", level='warning')
            # Continue to next method if failed
    
    # Method 3: Return default message for password-protected files or when other methods fail
    if has_password:
        return "Password protected file, cannot get author info via openpyxl"
    else:
        return "Cannot get author info"
"""
def get_workbook_essential_properties(workbook_obj):
    """
    Get essential properties from opened Excel workbook object (only show Last Author and Last Save Time)
    Return dictionary containing simplified workbook metadata
    """
    properties = {}  # Initialize properties dictionary
    
    try:
        # Get built-in document properties
        builtin_props = workbook_obj.BuiltinDocumentProperties  # Get built-in properties collection
        
        # Only process the two required properties
        try:
            last_author = builtin_props("Last Author").Value  # Get last author property value
            if last_author is not None:
                properties["👤 Last Author"] = str(last_author)  # Convert to string for display consistency
            else:
                properties["👤 Last Author"] = "Not set"  # Default message for empty property
        except Exception:
            properties["👤 Last Author"] = "Unable to retrieve"  # Error message for inaccessible property
        
        try:
            last_save_time = builtin_props("Last Save Time").Value  # Get last save time property value
            if last_save_time is not None:
                # Format datetime object for better display
                if isinstance(last_save_time, datetime):
                    last_save_time = last_save_time.strftime('%Y-%m-%d %H:%M:%S')
                properties["🕒 Last Save Time"] = str(last_save_time)  # Convert to string for display consistency
            else:
                properties["🕒 Last Save Time"] = "Not set"  # Default message for empty property
        except Exception:
            properties["🕒 Last Save Time"] = "Unable to retrieve"  # Error message for inaccessible property
            
    except Exception as e:
        console_print(f"Error getting workbook properties: {str(e)}", level='warning')
        properties["⚠️ Error"] = "Failed to retrieve workbook properties"
    
    return properties

def get_file_last_save_time(file_path):
    """Get file's last save time at OS level"""
    if not os.path.exists(file_path):
        return None

    try:
        timestamp = os.path.getmtime(file_path)  # Get file modification timestamp 
        last_save_time = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
        return last_save_time
    except Exception as e:
        console_print(f"Cannot get last save time for file '{file_path}': {str(e)}", level='warning')
        return None

def is_excel_file_accessible(file_path, open_password=None):
    """Check if Excel file is accessible"""
    if open_password is not None:
        console_print(f"File has password protection, skipping accessibility check")
        return True
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)  # Try to open file
        wb.close()  # Close file
        return True
    except Exception as e:
        console_print(f"File '{file_path}' is not accessible: {str(e)}", level='error')
        return False

def refresh_workbook_connections(workbook):
    """Refresh all links and connections in workbook (improved version: update each external link individually)"""
    refresh_count = 0  # Record refresh count
    
    try:
        # Handle Excel file links
        excel_links = workbook.LinkSources(Type=win32.constants.xlExcelLinks)
        if excel_links:
            console_print(f"📊 Found {len(excel_links)} Excel file links")
            for i, link_path in enumerate(excel_links):
                link_file_name = os.path.basename(link_path)  # Get linked file name
                link_dir = os.path.dirname(link_path)  # Get linked file directory
                console_print(f"  └─ Link {i+1}: {link_file_name}")
                console_print(f"     📁 Path: {link_dir}")
                
                if os.path.exists(link_path):  # Check if linked file exists
                    # Check if linked file has password (simplified handling, assume no password)
                    link_author = get_last_save_author_improved(link_path, False)
                    link_last_save_time = get_file_last_save_time(link_path)
                    console_print(f"     👤 Last author: {link_author}")
                    console_print(f"     🕒 Last save time: {link_last_save_time}")
                    
                    # 🔄 Added: Update each Excel file link individually
                    try:
                        console_print(f"     🔄 Updating link...")
                        workbook.UpdateLink(Name=link_path, Type=win32.constants.xlExcelLinks)
                        console_print(f"     ✅ Link update successful: {link_file_name}")
                        refresh_count += 1
                    except Exception as e:
                        console_print(f"     ❌ Link update failed: {str(e)}", level='error')
                        refresh_count += 1  # Count as processed but failed
                        
                else:
                    console_print(f"     ⚠️ Linked file does not exist: {link_path}", level='warning')
                    console_print(f"     ⏭️ Skipping this link update")
                
            console_print("")
        
        # Handle other data connections
        connections = workbook.Connections
        if connections and connections.Count > 0:
            console_print(f"🔗 Found {connections.Count} data connections")
            for i in range(1, connections.Count + 1):
                try:
                    connection = connections.Item(i)  # Get connection object
                    connection_name = connection.Name  # Get connection name
                    console_print(f"  └─ Connection {i}: {connection_name}")
                    
                    # Safely refresh connection
                    console_print(f"     🔄 Refreshing connection...")
                    safe_execute(connection.Refresh)  # Refresh connection
                    time.sleep(1)  # Delay between operations
                    refresh_count += 1
                    console_print(f"     ✅ Refresh completed")
                except Exception as e:
                    console_print(f"     ❌ Refresh failed: {str(e)}", level='error')
            console_print("")
        
        # Force recalculate all formulas
        if USER_CONFIG["advanced_settings"]["force_calculation"]:
            console_print("🧮 Performing full formula recalculation...")
            workbook.Application.CalculateFullRebuild()  # Force full rebuild calculation
            console_print("   ✅ Formula calculation completed")
            console_print("")
        
    except Exception as e:
        console_print(f"Error occurred while refreshing links: {str(e)}", level='error')
        raise ExcelAutomationError(f"Cannot refresh workbook links: {str(e)}")
    
    return refresh_count

def execute_macro_safely(excel_app, macro_name):
    """Safely execute macro"""
    try:
        console_print(f"⚙️ Attempting to execute macro: {macro_name}")
        safe_execute(excel_app.Run, macro_name)  # Execute macro
        console_print(f"   ✅ Macro execution successful")
        return True
    except Exception as e:
        console_print(f"   ❌ Macro execution failed: {str(e)}", level='error')
        return False

def get_last_save_author_improved(file_path, has_password=False, open_password=None, workbook_obj=None):
    """
    Improved version: Get Excel file's last author information using multiple methods
    Priority: 1. From opened workbook object 2. From openpyxl 3. Return default message
    For password-protected files, use win32com to open and extract information
    """
    # Method 1: If workbook object is provided, get author from opened workbook
    if workbook_obj is not None:
        try:
            # Get author information from opened Excel workbook object
            builtin_props = workbook_obj.BuiltinDocumentProperties  # Get built-in document properties
            last_author = builtin_props("Last Author").Value  # Get last author property
            return last_author if last_author else "Last author info not found from opened workbook"
        except Exception as e:
            console_print(f"Cannot get author info from opened workbook: {str(e)}", level='warning')
            # Continue to next method if failed
    
    # Method 2: For password-protected files, use win32com to open directly and extract
    if has_password and open_password is not None:
        excel_app = None  # Excel application object
        workbook = None   # Workbook object
        try:
            console_print(f"Using win32com to open password-protected file for metadata extraction...")
            
            # Start Excel application (background mode)
            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False  # Run in background
            excel_app.DisplayAlerts = False  # Don't show alerts
            excel_app.EnableEvents = False   # Disable events
            
            # Open workbook (read-only mode)
            open_params = {
                'Filename': file_path,
                'Password': open_password,
                'ReadOnly': True,  # Read-only mode to avoid file locking
                'UpdateLinks': False,  # Don't update links to improve opening speed
                'IgnoreReadOnlyRecommended': True
            }
            
            workbook = excel_app.Workbooks.Open(**open_params)
            
            # Get author information
            builtin_props = workbook.BuiltinDocumentProperties
            last_author = builtin_props("Last Author").Value
            
            return last_author if last_author else "Last author info not found in password-protected file"
            
        except Exception as e:
            console_print(f"Failed to open password file with win32com: {str(e)}", level='warning')
            return f"Password-protected file, cannot extract author info: {str(e)}"
        finally:
            # Clean up resources
            if workbook:
                try:
                    workbook.Close(SaveChanges=False)
                except:
                    pass
            if excel_app:
                try:
                    excel_app.Quit()
                except:
                    pass
    
    # Method 3: Use openpyxl for files without password
    if not has_password:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)  # Open workbook in read-only mode
            last_author = wb.properties.lastModifiedBy  # Get last modified by
            wb.close()  # Close workbook
            return last_author if last_author else "Last author info not found from openpyxl"
        except Exception as e:
            console_print(f"Cannot get author info via openpyxl: {str(e)}", level='warning')
            # Continue to next method if failed
    
    # Method 4: Return default message
    if has_password:
        return "Password-protected file, openpyxl cannot get author info"
    else:
        return "Cannot get author info"

def get_workbook_metadata_via_win32com(file_path, open_password=None, write_password=None):
    """
    New function: Use win32com to directly open file and extract complete metadata
    Specifically designed for metadata extraction from password-protected files
    Return dictionary containing Last Author and Last Save Time
    """
    excel_app = None  # Excel application object
    workbook = None   # Workbook object
    metadata = {}     # Metadata dictionary
    
    try:
        console_print(f"Using win32com to extract file metadata...")
        
        # Start Excel application (background mode)
        excel_app = win32.Dispatch("Excel.Application")
        excel_app.Visible = False  # Run in background without showing Excel window
        excel_app.DisplayAlerts = False  # Don't show alert dialogs
        excel_app.EnableEvents = False   # Disable event handling for better performance
        
        # Prepare opening parameters
        open_params = {
            'Filename': file_path,           # File path
            'ReadOnly': True,                # Read-only mode to avoid file locking
            'UpdateLinks': False,            # Don't update links for faster opening
            'IgnoreReadOnlyRecommended': True # Ignore read-only recommendation
        }
        
        # Add password parameters if available
        if open_password:
            open_params['Password'] = open_password
        if write_password:
            open_params['WriteResPassword'] = write_password
        
        # Open workbook
        workbook = excel_app.Workbooks.Open(**open_params)
        
        # Get built-in document properties
        builtin_props = workbook.BuiltinDocumentProperties
        
        # Extract Last Author
        try:
            last_author = builtin_props("Last Author").Value
            if last_author is not None:
                metadata["👤 Last Author"] = str(last_author)
            else:
                metadata["👤 Last Author"] = "Not set"
        except Exception:
            metadata["👤 Last Author"] = "Unable to retrieve"
        
        # Extract Last Save Time
        try:
            last_save_time = builtin_props("Last Save Time").Value
            if last_save_time is not None:
                # Format datetime object
                if isinstance(last_save_time, datetime):
                    last_save_time = last_save_time.strftime('%Y-%m-%d %H:%M:%S')
                metadata["🕒 Last Save Time"] = str(last_save_time)
            else:
                metadata["🕒 Last Save Time"] = "Not set"
        except Exception:
            metadata["🕒 Last Save Time"] = "Unable to retrieve"
        
        console_print(f"Successfully extracted metadata:")
        console_print(f"   👤 Last author: {metadata.get('👤 Last Author', 'N/A')}")
        console_print(f"   🕒 Last save time: {metadata.get('🕒 Last Save Time', 'N/A')}")
        
        return metadata
        
    except Exception as e:
        console_print(f"Failed to extract metadata using win32com: {str(e)}", level='error')
        return {
            "👤 Last Author": f"Extraction failed: {str(e)}",
            "🕒 Last Save Time": f"Extraction failed: {str(e)}"
        }
    finally:
        # Clean up resources
        if workbook:
            try:
                workbook.Close(SaveChanges=False)  # Close workbook without saving changes
            except Exception as e:
                console_print(f"Error occurred while closing workbook: {str(e)}", level='warning')
        
        if excel_app:
            try:
                excel_app.EnableEvents = True  # Re-enable events
                excel_app.Quit()               # Close Excel application
            except Exception as e:  
                console_print(f"Error occurred while closing Excel application: {str(e)}", level='warning')

def automate_excel_refresh_links(excel_file_path, file_config):
    """
    Modified version: Process single Excel file, optimized metadata extraction for password-protected files
    Avoid repeatedly opening files, directly use win32com to get complete metadata
    """
    macro_to_run = file_config.get("macro")  # Get the macro name to execute
    file_open_password = file_config.get("open_password")  # Get the open password
    file_write_password = file_config.get("write_password")  # Get the write password
    
    file_name = os.path.basename(excel_file_path)  # Get the file name
    console_print("")
    console_print(f"📁 Processing file: {file_name}")
    console_print("=" * 60)
    
    if not os.path.exists(excel_file_path):
        console_print(f"❌ File does not exist: {file_name}", level='error')
        return False
    
    # Check if file is accessible (only check for non-password files)
    if not is_excel_file_accessible(excel_file_path, file_open_password):
        console_print(f"❌ File is not accessible, skipping processing: {file_name}", level='error')
        return False
    
    excel_app = None  # Excel application object
    workbook = None   # Workbook object
    success = False   # Processing success flag
    has_password = file_open_password is not None  # Check if password exists
    
    # Initialize metadata variables
    metadata_before = {}  # Metadata before processing
    metadata_after = {}   # Metadata after processing
    
    try:
        # 🆕 For password-protected files, use win32com to get complete metadata before processing
        if has_password:
            console_print("📋 Getting metadata for password-protected file before processing...")
            metadata_before = get_workbook_metadata_via_win32com(
                excel_file_path, 
                file_open_password, 
                file_write_password
            )
        else:
            # For non-password files, use original method
            console_print("📋 Getting file metadata before processing...")
            last_save_time_before = get_file_last_save_time(excel_file_path)
            last_author_before = get_last_save_author_improved(
                excel_file_path, 
                has_password, 
                file_open_password
            )
            metadata_before = {
                "👤 Last Author": last_author_before,
                "🕒 Last Save Time": last_save_time_before
            }
            console_print(f"   🕒 Last save time before processing: {last_save_time_before}")
            console_print(f"   👤 Last author before processing: {last_author_before}")
        
        console_print("")
        
        # Start Excel application for actual processing
        console_print("🚀 Starting Excel application for processing...")
        excel_app = win32.Dispatch("Excel.Application")  # Create Excel application object
        excel_app.Visible = USER_CONFIG["advanced_settings"]["excel_visible"]  # Set visibility
        excel_app.DisplayAlerts = False  # Don't display alert messages
        excel_app.EnableEvents = False   # Disable events to improve performance
        console_print("   ✅ Excel application startup completed")
        
        # Open workbook for processing
        console_print(f"📂 Opening file for processing: {file_name}")
        
        # Use different opening methods based on password settings
        open_params = {
            'Filename': excel_file_path,         # File path
            'UpdateLinks': 3,                    # Update links
            'ReadOnly': False,                   # Non-read-only mode
            'IgnoreReadOnlyRecommended': True,   # Ignore read-only recommendation
            'Origin': win32.constants.xlWindows  # Origin is Windows
        }
        
        if file_open_password:  # If there's an open password
            open_params['Password'] = file_open_password
        if file_write_password:  # If there's a write password
            open_params['WriteResPassword'] = file_write_password
            
        workbook = safe_execute(excel_app.Workbooks.Open, **open_params)  # Open workbook
            
        console_print(f"   ✅ File opened successfully for processing")
        console_print("")
        
        # Refresh all links and connections
        console_print("🔄 Starting to refresh all external links and data connections...")
        refresh_count = refresh_workbook_connections(workbook)
        console_print(f"✅ Total refreshed {refresh_count} links/connections")
        console_print("")
        
        # Execute macro (if specified)
        if macro_to_run:
            execute_macro_safely(excel_app, macro_to_run)
        else:
            console_print("ℹ️ No macro specified to execute")
        console_print("")
        
        # Save file
        console_print(f"💾 Saving file: {file_name}")
        
        # Check workbook status
        console_print(f"   📊 Workbook modification status: {'Modified' if not workbook.Saved else 'Not modified'}")
        console_print(f"   🔒 Workbook read/write status: {'Read-only' if workbook.ReadOnly else 'Writable'}")
        
        safe_execute(workbook.Save)  # Save workbook
        console_print(f"   ✅ File saved successfully")
        console_print("")
        
        # 🆕 Get metadata after processing (unified from opened workbook)
        console_print("📋 Getting final file metadata after processing...")
        
        # Get metadata after processing from opened workbook
        try:
            builtin_props = workbook.BuiltinDocumentProperties
            
            # Get Last Author
            try:
                last_author_after = builtin_props("Last Author").Value
                if last_author_after is not None:
                    metadata_after["👤 Last Author"] = str(last_author_after)
                else:
                    metadata_after["👤 Last Author"] = "Not set"
            except Exception:
                metadata_after["👤 Last Author"] = "Unable to retrieve"
            
            # Get file system level last save time (more accurate)
            last_save_time_after = get_file_last_save_time(excel_file_path)
            metadata_after["🕒 Last Save Time"] = last_save_time_after
            
        except Exception as e:
            console_print(f"Error occurred while getting metadata after processing: {str(e)}", level='warning')
            metadata_after = {
                "👤 Last Author": "Retrieval failed",
                "🕒 Last Save Time": get_file_last_save_time(excel_file_path)
            }
        
        console_print(f"   👤 Last author after processing: {metadata_after.get('👤 Last Author')}")
        console_print(f"   🕒 Last save time after processing: {metadata_after.get('🕒 Last Save Time')}")
        
        # 🆕 Perform metadata comparison
        console_print("📊 Metadata comparison:")
        
        # Compare author information
        author_before = metadata_before.get("👤 Last Author")
        author_after = metadata_after.get("👤 Last Author")
        if author_before and author_after:
            if author_before != author_after:
                console_print(f"   👤 Author changed: {author_before} → {author_after}")
            else:
                console_print(f"   👤 Author unchanged: {author_after}")
        else:
            console_print(f"   👤 Author after processing: {author_after}")
        
        # Compare save time
        time_before = metadata_before.get("🕒 Last Save Time")
        time_after = metadata_after.get("🕒 Last Save Time")
        if time_before and time_after:
            if time_before != time_after:
                console_print(f"   🕒 Save time updated: {time_before} → {time_after}")
            else:
                console_print(f"   🕒 Save time unchanged: {time_after}")
        else:
            console_print(f"   🕒 Save time after processing: {time_after}")
        
        success = True  # Mark processing as successful
        
    except Exception as e:
        console_print(f"❌ Error occurred while processing file: {str(e)}", level='error')
        success = False
        
    finally:
        # Clean up resources
        if workbook:
            try:
                workbook.Close(SaveChanges=False)  # Close workbook (don't save changes)
                console_print("🔐 Workbook closed")
            except Exception as e:
                console_print(f"⚠️ Error occurred while closing workbook: {str(e)}", level='warning')
                
        if excel_app:
            try:
                excel_app.EnableEvents = True  # Re-enable events
                excel_app.Quit()               # Quit Excel application
                console_print("🔐 Excel application closed")
            except Exception as e:
                console_print(f"⚠️ Error occurred while closing Excel application: {str(e)}", level='warning')
    
    status_icon = "✅" if success else "❌"  # Choose icon based on success or failure
    console_print(f"{status_icon} File '{file_name}' processing {'successful' if success else 'failed'}")
    console_print("=" * 60)
    return success

def process_excel_files_in_directory(base_directory, file_configs):
    """
    Process multiple Excel files in specified directory, following the order of FILE_CONFIGS keys
    """
    console_print("")
    console_print(f"🚀 Starting batch processing directory: {base_directory}")
    console_print(f"⏰ Processing start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    console_print("=" * 70)
    
    if not os.path.isdir(base_directory):
        console_print(f"❌ Directory does not exist: {base_directory}", level='error')
        return
    
    processed_files = []  # Successfully processed files list
    failed_files = []    # Failed processing files list
    skipped_files = []   # Skipped files list
    
    # Get all Excel files in the directory
    all_excel_files = [f for f in os.listdir(base_directory) 
                      if f.lower().endswith(('.xlsx', '.xlsm')) and 
                      os.path.isfile(os.path.join(base_directory, f))]
    
    console_print(f"📊 Found {len(all_excel_files)} Excel files in directory")
    console_print("")
    
    # Process files in the order of FILE_CONFIGS keys
    for prefix in file_configs.keys():  # Iterate through FILE_CONFIGS keys in order
        console_print(f"🔍 Searching for files with prefix: {prefix}")
        matched_files = [f for f in all_excel_files if f.startswith(prefix)]  # Find matching files
        
        if not matched_files:
            console_print(f"⚠️ No files found with prefix: {prefix}", level='warning')
            skipped_files.append(f"No files for prefix: {prefix}")
            continue
        
        for filename in matched_files:
            full_file_path = os.path.join(base_directory, filename)  # Full file path
            console_print(f"🎯 Found matching file: {filename} (prefix: '{prefix}')")
            current_file_config = file_configs[prefix]  # Get corresponding configuration
            
            if automate_excel_refresh_links(full_file_path, current_file_config):
                processed_files.append(filename)  # Add to success list
            else:
                failed_files.append(filename)     # Add to failed list
    
    # Check for files that don't match any prefix
    for filename in all_excel_files:
        if not any(filename in processed_files or filename in failed_files for filename in all_excel_files):
            console_print(f"⏭️ Skipping file (no matching prefix): {filename}")
            skipped_files.append(filename)
    
    # Output processing summary
    console_print("")
    console_print("=" * 70)
    console_print("📊 Batch processing completion summary:")
    console_print(f"   ✅ Successfully processed: {len(processed_files)} files")
    console_print(f"   ❌ Processing failed: {len(failed_files)} files")
    console_print(f"   ⏭️ Skipped files: {len(skipped_files)} files")
    console_print("")
    
    if processed_files:
        console_print("✅ Successfully processed files:")
        for file in processed_files:
            console_print(f"   • {file}")
        console_print("")
    
    if failed_files:
        console_print("❌ Processing failed files:")
        for file in failed_files:
            console_print(f"   • {file}")
        console_print("")
    
    if skipped_files:
        console_print("⏭️ Skipped files:")
        for file in skipped_files:
            console_print(f"   • {file}")
        console_print("")
    
    console_print(f"⏰ Processing end time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    console_print("=" * 70)

def validate_configuration():
    """Validate user configuration"""
    errors = []
    
    # Check base directory
    if not os.path.exists(USER_CONFIG["base_directory"]):
        errors.append(f"Base directory does not exist: {USER_CONFIG['base_directory']}")
    
    # Check file configurations
    if not USER_CONFIG["FILE_CONFIGS"]:
        errors.append("No file configurations specified")
    
    # Check advanced settings
    advanced = USER_CONFIG["advanced_settings"]
    if advanced["max_retries"] < 1:
        errors.append("max_retries must be at least 1")
    if advanced["retry_delay_base"] < 1:
        errors.append("retry_delay_base must be at least 1")
    
    return errors

def main():
    global logger
    log_filepath = None

    try:
        config_errors = validate_configuration()
        if config_errors:
            print("❌ Configuration errors found:")
            for error in config_errors:
                print(f"   • {error}")
            return 1

        logger = setup_logging()
        log_filepath = os.environ.get("log_filepath")

        process_excel_files_in_directory(
            USER_CONFIG["base_directory"],
            USER_CONFIG["FILE_CONFIGS"]
        )

        console_print("")
        console_print("🎉 Program execution completed")
        console_print("="*80)
        return 0

    except KeyboardInterrupt:
        console_print("")
        console_print("⏹️ Program interrupted by user", level='warning')
        return 1
    except Exception as e:
        console_print("")
        console_print(f"💥 Unexpected error occurred during program execution: {str(e)}", level='error')
        return 1
    finally:
        if logger and log_filepath and os.path.exists(log_filepath):
            console_print("Preparing to send notification email...")

            content = ""
            try:
                with open(log_filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
            except Exception as e:
                content = f"Could not read log file: {e}"

            email_body = f"""Hello Team,
                        This is an automated notification.

                        {content}

                        Best regards,
                        Your Automation Script
                        """
            send_outlook_email(
                to_recipients=TO_RECIPIENTS,
                subject=f"{EMAIL_SUBJECT} ({time.strftime('%Y-%m-%d %H:%M:%S')})",
                body=email_body,
                cc_recipients=CC_RECIPIENTS,
                bcc_recipients=BCC_RECIPIENTS
            )
            console_print("📄 Notification email sent.")
        elif not log_filepath:
            print("Log file path not found, cannot send email.")		


    finally:
        if logger:
            console_print("📄 Log file saved successfully")

if __name__ == "__main__":
    main()





















































